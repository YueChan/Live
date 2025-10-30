#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Raw Data矩阵统计分析工具 V2.0
全新独立版本 - 统计每个BIN的Raw Data矩阵（AVG/MAX/MIN/STD）
"""

import os
import pandas as pd
import numpy as np
from pathlib import Path
from tkinter import Tk, filedialog, messagebox
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# 文件选择和基础函数
# ============================================================================

def select_folder():
    """选择文件夹"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    print("\n" + "="*80)
    print("步骤 1/4: 选择包含CSV日志文件的文件夹".center(80))
    print("="*80)

    folder_path = filedialog.askdirectory(title="请选择包含CSV日志文件的文件夹")
    root.destroy()

    if not folder_path:
        print("❌ 未选择文件夹")
        return None

    print(f"\n✅ 已选择: {folder_path}")
    return folder_path

def select_save_location():
    """选择保存位置"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_name = f"RawData_Stats_{timestamp}.xlsx"

    file_path = filedialog.asksaveasfilename(
        title="选择保存位置",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    root.destroy()

    if not file_path:
        print("❌ 未选择保存位置")
        return None

    print(f"\n✅ 将保存到: {file_path}")
    return file_path

def find_csv_files(folder_path):
    """查找CSV文件"""
    print("\n" + "="*80)
    print("步骤 2/4: 扫描CSV文件".center(80))
    print("="*80)

    csv_files = list(Path(folder_path).glob("*.csv"))

    if not csv_files:
        print(f"\n❌ 未找到CSV文件")
        return []

    print(f"\n✅ 找到 {len(csv_files)} 个CSV文件")
    return csv_files

# ============================================================================
# 核心解析函数 - 全新实现
# ============================================================================

def read_csv_file(file_path):
    """读取CSV文件 - 支持多种编码"""
    encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'latin1']

    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding, errors='ignore', newline='') as f:
                content = f.read()
            # 统一行尾
            content = content.replace('\r\n', '\n').replace('\r', '\n')
            lines = content.split('\n')
            return lines, encoding
        except:
            continue

    return None, None

def extract_raw_data_matrix(lines, start_line_idx):
    """从指定位置提取Raw Data矩阵

    参数:
        lines: 文件的所有行
        start_line_idx: "Raw Data:" 所在行的索引

    返回:
        matrix: 二维列表，每个元素是一行的数值列表
    """

    print(f"\n      🔍 开始提取矩阵 (从第{start_line_idx}行)")

    matrix = []
    i = start_line_idx + 1  # 跳过 "Raw Data:" 行

    # 检查并跳过标题行
    if i < len(lines) and 'Column' in lines[i]:
        print(f"      → 跳过标题行: {lines[i][:50]}...")
        i += 1

    # 显示接下来要解析的行
    print(f"      → 开始解析数据行...")

    parsed_count = 0
    while i < len(lines):
        line = lines[i]
        line_stripped = line.strip()

        # 停止条件
        if not line_stripped:
            break
        if any(kw in line_stripped for kw in ['Limits:', 'PASS/FAIL:', 'Test Started', 'Bin #:', 'END_UUT']):
            break

        # 解析数据行 "     0 : <数据>"
        if ':' in line:
            try:
                # 分割冒号
                colon_pos = line.index(':')
                row_part = line[:colon_pos].strip()
                data_part = line[colon_pos+1:]

                # 检查行号是否为数字
                try:
                    row_num = int(row_part)
                except:
                    i += 1
                    continue

                # 解析数据部分
                values = []

                # 检测分隔符
                if '\t' in data_part:
                    # 制表符分隔
                    parts = data_part.split('\t')
                elif ',' in data_part:
                    # 逗号分隔
                    parts = data_part.split(',')
                else:
                    # 空格分隔（使用split()自动处理多个空格）
                    parts = data_part.split()

                # 转换为数值
                for p in parts:
                    p = p.strip()
                    if p:
                        try:
                            values.append(float(p))
                        except:
                            pass

                # 调试第一行
                if parsed_count == 0 and values:
                    print(f"      ✅ 第一行 (Row {row_num}):")
                    print(f"         原始: {line[:100]}...")
                    print(f"         数据部分: {data_part[:100]}...")
                    print(f"         检测分隔符: Tab={chr(9) in data_part}, 逗号={',' in data_part}")
                    print(f"         解析结果: {len(values)} 个数值")
                    if len(values) >= 5:
                        print(f"         前5个: {values[:5]}")
                        print(f"         后5个: {values[-5:]}")
                    else:
                        print(f"         所有值: {values}")

                if values:
                    matrix.append(values)
                    parsed_count += 1

            except Exception as e:
                if parsed_count == 0:
                    print(f"      ❌ 解析错误: {e}")

        i += 1

    print(f"      ✅ 提取完成: {len(matrix)} 行")
    if matrix:
        print(f"         矩阵尺寸: {len(matrix)} × {len(matrix[0])}")

    return matrix

def parse_one_csv_file(file_path, file_name):
    """解析一个CSV文件，提取所有BIN的Raw Data"""

    lines, encoding = read_csv_file(file_path)
    if lines is None:
        return None, None

    results = []

    # 查找所有的 "Bin #:" 和对应的 "Raw Data:"
    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # 找到 Bin #
        if line.startswith('Bin #:'):
            bin_num = line.split(':', 1)[1].strip()

            # 向下查找测试名称和Raw Data
            test_name = None
            raw_data_idx = None

            for j in range(i+1, min(i+300, len(lines))):
                if lines[j].strip().startswith('Name:'):
                    test_name = lines[j].strip().split(':', 1)[1].strip()
                elif lines[j].strip() == 'Raw Data:':
                    raw_data_idx = j
                    break
                elif lines[j].strip().startswith('Bin #:'):
                    break

            # 如果找到了Raw Data，提取矩阵
            if raw_data_idx is not None:
                matrix = extract_raw_data_matrix(lines, raw_data_idx)

                if matrix and len(matrix) > 0:
                    results.append({
                        'bin_num': bin_num,
                        'test_name': test_name or 'Unknown',
                        'matrix': matrix,
                        'source_file': file_name
                    })

            i = j if raw_data_idx else i + 1
        else:
            i += 1

    return results, encoding

def parse_all_files(csv_files):
    """解析所有CSV文件"""
    print("\n" + "="*80)
    print("步骤 3/4: 解析CSV文件".center(80))
    print("="*80)

    all_results = []

    for idx, csv_file in enumerate(csv_files, 1):
        print(f"\n[{idx}/{len(csv_files)}] {csv_file.name}")

        try:
            results, encoding = parse_one_csv_file(csv_file, csv_file.name)

            if results:
                print(f"   ✅ 找到 {len(results)} 个BIN测试")
                for r in results:
                    rows = len(r['matrix'])
                    cols = len(r['matrix'][0]) if rows > 0 else 0
                    print(f"      • Bin #{r['bin_num']}: {r['test_name'][:40]} - {rows}×{cols}")

                all_results.extend(results)
            else:
                print(f"   ⚠️  未找到Raw Data")

        except Exception as e:
            print(f"   ❌ 解析失败: {e}")

    return all_results

# ============================================================================
# 统计计算
# ============================================================================

def calculate_statistics(all_results):
    """按BIN编号分组，计算每个单元格的统计数据"""
    print("\n" + "="*80)
    print("计算统计数据".center(80))
    print("="*80)

    # 按BIN分组
    bin_groups = {}
    for result in all_results:
        bin_num = result['bin_num']
        if bin_num not in bin_groups:
            bin_groups[bin_num] = {
                'test_name': result['test_name'],
                'matrices': []
            }
        bin_groups[bin_num]['matrices'].append(result['matrix'])

    # 计算统计
    statistics = {}

    for bin_num in sorted(bin_groups.keys(), key=lambda x: (isinstance(x, str), x)):
        print(f"\nBin #{bin_num}:")

        matrices = bin_groups[bin_num]['matrices']
        test_name = bin_groups[bin_num]['test_name']

        # 找最小尺寸
        shapes = [(len(m), len(m[0]) if m else 0) for m in matrices]
        min_rows = min(s[0] for s in shapes)
        min_cols = min(s[1] for s in shapes)

        print(f"   样本数: {len(matrices)}, 尺寸: {min_rows}×{min_cols}")

        # 初始化统计矩阵
        avg_mat = np.zeros((min_rows, min_cols))
        max_mat = np.zeros((min_rows, min_cols))
        min_mat = np.zeros((min_rows, min_cols))
        std_mat = np.zeros((min_rows, min_cols))

        # 计算每个单元格
        for r in range(min_rows):
            for c in range(min_cols):
                values = [m[r][c] for m in matrices if r < len(m) and c < len(m[r])]
                if values:
                    avg_mat[r, c] = np.mean(values)
                    max_mat[r, c] = np.max(values)
                    min_mat[r, c] = np.min(values)
                    std_mat[r, c] = np.std(values, ddof=1) if len(values) > 1 else 0

        statistics[bin_num] = {
            'test_name': test_name,
            'sample_count': len(matrices),
            'shape': (min_rows, min_cols),
            'avg': avg_mat,
            'max': max_mat,
            'min': min_mat,
            'std': std_mat
        }

    print(f"\n✅ 完成统计: {len(statistics)} 个BIN")
    return statistics

# ============================================================================
# 保存Excel
# ============================================================================

def save_to_excel(statistics, output_path):
    """保存到Excel - 每个BIN一个工作表，纵向排列AVG/MAX/MIN/STD"""
    print("\n" + "="*80)
    print("步骤 4/4: 保存Excel文件".center(80))
    print("="*80)

    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 汇总信息
            summary_data = {
                '项目': ['分析时间', 'BIN数量', '生成工作表'],
                '详情': [
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    len(statistics),
                    len(statistics) + 1
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='汇总', index=False)

            # 每个BIN一个工作表
            for bin_num, data in sorted(statistics.items(), key=lambda x: (isinstance(x[0], str), x[0])):
                rows, cols = data['shape']
                test_name = data['test_name']

                # 创建工作表
                sheet_name = f"Bin{bin_num}"[:31]

                # 创建一个空的DataFrame用于布局
                # 纵向排列：AVG部分 -> 空2行 -> MAX部分 -> 空2行 -> MIN部分 -> 空2行 -> STD部分
                all_rows = []

                # 添加标题和测试名称
                all_rows.append(['', f'Bin #{bin_num}: {test_name}'])
                all_rows.append([''])

                # AVG部分
                all_rows.append(['===== 平均值 (AVG) ====='])
                avg_header = ['Row'] + [f'Col{i}' for i in range(cols)]
                all_rows.append(avg_header)
                for r in range(rows):
                    row_data = [r] + list(data['avg'][r, :])
                    all_rows.append(row_data)

                # 空2行
                all_rows.append([''])
                all_rows.append([''])

                # MAX部分
                all_rows.append(['===== 最大值 (MAX) ====='])
                max_header = ['Row'] + [f'Col{i}' for i in range(cols)]
                all_rows.append(max_header)
                for r in range(rows):
                    row_data = [r] + list(data['max'][r, :])
                    all_rows.append(row_data)

                # 空2行
                all_rows.append([''])
                all_rows.append([''])

                # MIN部分
                all_rows.append(['===== 最小值 (MIN) ====='])
                min_header = ['Row'] + [f'Col{i}' for i in range(cols)]
                all_rows.append(min_header)
                for r in range(rows):
                    row_data = [r] + list(data['min'][r, :])
                    all_rows.append(row_data)

                # 空2行
                all_rows.append([''])
                all_rows.append([''])

                # STD部分
                all_rows.append(['===== 标准差 (STD) ====='])
                std_header = ['Row'] + [f'Col{i}' for i in range(cols)]
                all_rows.append(std_header)
                for r in range(rows):
                    row_data = [r] + list(data['std'][r, :])
                    all_rows.append(row_data)

                # 转换为DataFrame并写入
                df = pd.DataFrame(all_rows)
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                print(f"   • {sheet_name}: {test_name[:40]}")

                # 格式化工作表
                ws = writer.sheets[sheet_name]

                # 计算各部分的起始行（Excel行号从1开始）
                title_row = 1
                avg_title_row = 3
                avg_header_row = 4
                avg_data_start = 5
                avg_data_end = avg_data_start + rows - 1

                max_title_row = avg_data_end + 3
                max_header_row = max_title_row + 1
                max_data_start = max_header_row + 1
                max_data_end = max_data_start + rows - 1

                min_title_row = max_data_end + 3
                min_header_row = min_title_row + 1
                min_data_start = min_header_row + 1
                min_data_end = min_data_start + rows - 1

                std_title_row = min_data_end + 3
                std_header_row = std_title_row + 1
                std_data_start = std_header_row + 1
                std_data_end = std_data_start + rows - 1

                # 格式化标题行
                title_cell = ws.cell(title_row, 1)
                title_cell.font = Font(bold=True, size=12)

                # 格式化各部分标题
                for row_num, color in [(avg_title_row, "4472C4"), (max_title_row, "FF0000"),
                                       (min_title_row, "00B050"), (std_title_row, "FFC000")]:
                    cell = ws.cell(row_num, 1)
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                # 格式化表头行
                for row_num, color in [(avg_header_row, "4472C4"), (max_header_row, "FF0000"),
                                       (min_header_row, "00B050"), (std_header_row, "FFC000")]:
                    for col_idx in range(1, cols + 2):
                        cell = ws.cell(row_num, col_idx)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # 设置列宽
                ws.column_dimensions['A'].width = 8
                for col_idx in range(2, min(cols + 2, 40)):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 10

                # 数值格式和对齐
                for row_num in range(avg_data_start, std_data_end + 1):
                    for col_idx in range(1, cols + 2):
                        cell = ws.cell(row_num, col_idx)
                        if col_idx == 1:  # Row列
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:  # 数值列
                            if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                                cell.number_format = '0.00'
                                cell.alignment = Alignment(horizontal='right', vertical='center')

        print(f"\n✅ 保存成功: {output_path}")
        return True

    except Exception as e:
        print(f"\n❌ 保存失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def show_result_message(output_path, statistics):
    """显示完成消息"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    message = f"""✅ 分析完成！

BIN测试项目: {len(statistics)}
生成工作表: {len(statistics) * 4 + 1}

文件已保存:
{os.path.basename(output_path)}

保存目录:
{os.path.dirname(output_path)}
"""

    messagebox.showinfo("完成", message)
    root.destroy()

# ============================================================================
# 主程序
# ============================================================================

def main():
    print("""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║                   Raw Data 矩阵统计分析工具 V2.0                              ║
║                                                                               ║
║                         全新独立版本                                          ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
""")

    # 1. 选择文件夹
    folder_path = select_folder()
    if not folder_path:
        input("\n按回车退出...")
        return

    # 2. 查找CSV文件
    csv_files = find_csv_files(folder_path)
    if not csv_files:
        input("\n按回车退出...")
        return

    # 3. 解析所有文件
    all_results = parse_all_files(csv_files)
    if not all_results:
        print("\n❌ 没有解析到任何数据")
        input("\n按回车退出...")
        return

    # 4. 计算统计
    statistics = calculate_statistics(all_results)
    if not statistics:
        print("\n❌ 统计计算失败")
        input("\n按回车退出...")
        return

    # 5. 选择保存位置
    output_path = select_save_location()
    if not output_path:
        input("\n按回车退出...")
        return

    # 6. 保存Excel
    success = save_to_excel(statistics, output_path)

    if success:
        show_result_message(output_path, statistics)
        print("\n" + "="*80)
        print("✅ 程序执行完成".center(80))
        print("="*80)

    input("\n按回车退出...")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  用户中断")
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        input("\n按回车退出...")
